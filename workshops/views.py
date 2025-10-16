from datetime import date
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.core.exceptions import PermissionDenied
from django.db.models import Count, Q
from django.http import FileResponse, HttpRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.views.generic import DetailView, ListView, CreateView, UpdateView, TemplateView

from .models import Workshop, WorkshopDocument, WorkshopPhoto
from .forms import WorkshopForm, WorkshopPhotoFormSet, WorkshopDocumentFormSet, LoginForm


class HomeView(TemplateView):
    template_name = 'workshops/home.html'


class WorkshopListView(ListView):
    model = Workshop
    template_name = 'workshops/workshop_list.html'
    context_object_name = 'workshops'
    paginate_by = 10

    def get_queryset(self):
        qs = Workshop.objects.all().order_by('-start_date', '-created_at')
        tab = self.request.GET.get('tab', 'upcoming')
        today = date.today()
        if tab == 'live':
            qs = qs.filter(status=Workshop.Status.LIVE)
        elif tab == 'completed':
            qs = qs.filter(status=Workshop.Status.COMPLETED)
        else:
            qs = qs.filter(status=Workshop.Status.UPCOMING)

        topic = self.request.GET.get('topic')
        location = self.request.GET.get('location')
        mode = self.request.GET.get('mode')
        start = self.request.GET.get('start')
        end = self.request.GET.get('end')

        if topic:
            qs = qs.filter(Q(title__icontains=topic) | Q(topic__icontains=topic))
        if location:
            qs = qs.filter(Q(city__icontains=location) | Q(institute__icontains=location) | Q(state__icontains=location))
        if mode in [Workshop.Mode.PHYSICAL, Workshop.Mode.ONLINE]:
            qs = qs.filter(mode=mode)
        if start:
            qs = qs.filter(start_date__gte=start)
        if end:
            qs = qs.filter(end_date__lte=end)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['tab'] = self.request.GET.get('tab', 'upcoming')
        return ctx


class WorkshopDetailView(DetailView):
    model = Workshop
    template_name = 'workshops/workshop_detail.html'
    context_object_name = 'workshop'


class IsCoordinatorOrAdminMixin(UserPassesTestMixin):
    def test_func(self):
        workshop_id = self.kwargs.get('pk')
        if not workshop_id:
            return True
        workshop = get_object_or_404(Workshop, pk=workshop_id)
        user = self.request.user
        return user.is_superuser or workshop.coordinator_id == user.id


class WorkshopCreateView(LoginRequiredMixin, CreateView):
    model = Workshop
    form_class = WorkshopForm
    template_name = 'workshops/workshop_form.html'
    success_url = reverse_lazy('dashboard')

    def form_valid(self, form):
        form.instance.coordinator = self.request.user
        if not form.instance.coordinator_email:
            form.instance.coordinator_email = self.request.user.email
        messages.success(self.request, 'Workshop created successfully')
        return super().form_valid(form)


class WorkshopUpdateView(LoginRequiredMixin, IsCoordinatorOrAdminMixin, UpdateView):
    model = Workshop
    form_class = WorkshopForm
    template_name = 'workshops/workshop_form.html'
    success_url = reverse_lazy('dashboard')

    def handle_no_permission(self):
        raise PermissionDenied

    def form_valid(self, form):
        messages.success(self.request, 'Workshop updated successfully')
        return super().form_valid(form)


@login_required
def manage_uploads(request: HttpRequest, pk: int) -> HttpResponse:
    workshop = get_object_or_404(Workshop, pk=pk)
    if not (request.user.is_superuser or request.user.id == workshop.coordinator_id):
        raise PermissionDenied
    if request.method == 'POST':
        photo_formset = WorkshopPhotoFormSet(request.POST, request.FILES, instance=workshop, prefix='photos')
        doc_formset = WorkshopDocumentFormSet(request.POST, request.FILES, instance=workshop, prefix='docs')
        if photo_formset.is_valid() and doc_formset.is_valid():
            photo_formset.save()
            doc_formset.save()
            messages.success(request, 'Files updated successfully')
            return redirect('workshop-detail', pk=workshop.pk)
    else:
        photo_formset = WorkshopPhotoFormSet(instance=workshop, prefix='photos')
        doc_formset = WorkshopDocumentFormSet(instance=workshop, prefix='docs')
    return render(request, 'workshops/manage_uploads.html', {
        'workshop': workshop,
        'photo_formset': photo_formset,
        'doc_formset': doc_formset,
    })


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'workshops/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Year-wise stats
        year_counts = (
            Workshop.objects
            .values_list('start_date__year')
            .annotate(total=models.Count('id'))
            .order_by('start_date__year')
        )
        ctx['year_counts'] = year_counts
        ctx['total_workshops'] = Workshop.objects.count()
        ctx['total_participants'] = Workshop.objects.aggregate(total=models.Sum('participants_count'))['total'] or 0
        # Optional summaries
        ctx['state_counts'] = (
            Workshop.objects.exclude(state='')
            .values('state')
            .annotate(total=models.Count('id'))
            .order_by('-total')[:10]
        )
        ctx['category_counts'] = (
            Workshop.objects.exclude(category='')
            .values('category')
            .annotate(total=models.Count('id'))
            .order_by('-total')[:10]
        )
        if self.request.user.is_superuser:
            ctx['my_workshops'] = Workshop.objects.all().order_by('-created_at')[:20]
        else:
            ctx['my_workshops'] = Workshop.objects.filter(coordinator=self.request.user).order_by('-created_at')[:20]
        return ctx


def export_excel(request: HttpRequest) -> HttpResponse:
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Workshops'
    headers = [
        'Title', 'Topic', 'Start Date', 'End Date', 'Mode', 'City', 'Institute',
        'State', 'Coordinator', 'Email', 'Phone', 'Status', 'Participants', 'Category', 'Registration Link'
    ]
    ws.append(headers)
    for w in Workshop.objects.all().order_by('-start_date'):
        ws.append([
            w.title, w.topic, w.start_date, w.end_date, w.get_mode_display(), w.city, w.institute,
            w.state, w.coordinator.get_full_name() or w.coordinator.username,
            w.coordinator_email, w.coordinator_phone, w.get_status_display(),
            w.participants_count, w.category, w.registration_link,
        ])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return FileResponse(bio, as_attachment=True, filename='workshops.xlsx')


def export_pdf(request: HttpRequest) -> HttpResponse:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    bio = BytesIO()
    p = canvas.Canvas(bio, pagesize=A4)
    width, height = A4
    y = height - 50
    p.setFont('Helvetica-Bold', 14)
    p.drawString(50, y, 'Semiconductor EdTech Workshops Summary')
    y -= 30
    p.setFont('Helvetica', 10)
    for w in Workshop.objects.all().order_by('-start_date')[:100]:
        line = f"{w.start_date} - {w.title} ({w.get_status_display()}) - {w.city or w.state or w.institute}"
        p.drawString(50, y, line[:110])
        y -= 15
        if y < 50:
            p.showPage()
            y = height - 50
            p.setFont('Helvetica', 10)
    p.showPage()
    p.save()
    bio.seek(0)
    return FileResponse(bio, as_attachment=True, filename='workshops.pdf')


class LoginView(TemplateView):
    template_name = 'workshops/login.html'

    def post(self, request, *args, **kwargs):
        form = LoginForm(request.POST)
        if form.is_valid():
            identifier = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user_obj = User.objects.filter(Q(username=identifier) | Q(email__iexact=identifier)).first()
            username = user_obj.username if user_obj else identifier
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('dashboard')
        messages.error(request, 'Invalid credentials')
        return render(request, self.template_name, {'form': form})

    def get(self, request, *args, **kwargs):
        form = LoginForm()
        return render(request, self.template_name, {'form': form})


def logout_view(request):
    logout(request)
    return redirect('home')

# Create your views here.
